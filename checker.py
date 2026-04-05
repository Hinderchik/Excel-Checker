import requests
import csv
import openpyxl
from concurrent.futures import ThreadPoolExecutor
import argparse
import os

def check_url(url, timeout):
    if not url.startswith(('http://', 'https://')):
        url = 'https://' + url
    
    try:
        response = requests.get(url, timeout=timeout, allow_redirects=True)
        status = response.status_code
        final_url = response.url
        redirect_count = len(response.history)
        
        title = ''
        if '<title>' in response.text:
            start = response.text.find('<title>') + 7
            end = response.text.find('</title>')
            title = response.text[start:end][:80]
        
        return [url, status, title, final_url, redirect_count, '']
        
    except requests.exceptions.Timeout:
        return [url, 'ошибка', '', '', 0, 'таймаут']
    except Exception as e:
        return [url, 'ошибка', '', '', 0, str(e)[:40]]

def read_urls_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet = wb.active
    urls = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            urls.append(row[0])
    wb.close()
    return urls

def save_results(results, output_name):
    # CSV
    with open(f'{output_name}.csv', 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['исходный_url', 'статус', 'тайтл', 'финальный_url', 'число_редиректов', 'ошибка'])
        writer.writerows(results)
    
    # Excel
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(['исходный_url', 'статус', 'тайтл', 'финальный_url', 'число_редиректов', 'ошибка'])
    for row in results:
        sheet.append(row)
    wb.save(f'{output_name}.xlsx')
    
    # Битые ссылки отдельно
    broken = [row for row in results if row[1] in [404, 'ошибка']]
    if broken:
        wb_broken = openpyxl.Workbook()
        sheet_broken = wb_broken.active
        sheet_broken.append(['исходный_url', 'статус', 'тайтл', 'финальный_url', 'число_редиректов', 'ошибка'])
        for row in broken:
            sheet_broken.append(row)
        wb_broken.save(f'{output_name}_broken.xlsx')

parser = argparse.ArgumentParser(description='Массовая проверка сайтов')
parser.add_argument('input', nargs='?', default='links.xlsx', help='Входной Excel файл')
parser.add_argument('output', nargs='?', default='result', help='Имя выходного файла')
parser.add_argument('--timeout', type=int, default=10, help='Таймаут в секундах')
parser.add_argument('--workers', type=int, default=10, help='Количество потоков')
args = parser.parse_args()

print(f"Читаю файл: {args.input}")
urls = read_urls_from_excel(args.input)
print(f"Найдено ссылок: {len(urls)}")

with ThreadPoolExecutor(max_workers=args.workers) as executor:
    results = list(executor.map(lambda url: check_url(url, args.timeout), urls))

save_results(results, args.output)

working = len([r for r in results if r[1] == 200])
broken = len([r for r in results if r[1] == 404])
errors = len([r for r in results if r[1] == 'ошибка'])
redirects = len([r for r in results if r[4] > 0])

print("\nСтатистика:")
print(f"Работают: {working}")
print(f"Битые (404): {broken}")
print(f"Ошибки: {errors}")
print(f"С редиректами: {redirects}")
print(f"Результат: {args.output}.xlsx")